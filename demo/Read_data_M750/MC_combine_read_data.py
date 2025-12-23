# myarm_combined_logging.py - 读取myArmC角度并发送给myArmM执行，并采集MyArmM数据
import socket
import time
import threading
import serial.tools.list_ports
import os
import shutil
import traceback
from openpyxl import load_workbook, Workbook
from pymycobot import MyArmC, MyArmM
from pymycobot.error import MyArmDataException


def get_port():
    port_list = serial.tools.list_ports.comports()
    i = 1
    res = {}
    for port in port_list:
        print("{} - {}".format(i, port.device))
        res[str(i)] = port.device
        i += 1
    return res


def wait_for_button_press(c):
    print("等待按下红色按钮以启动摇操功能...")
    last_status = [0]
    while True:
        current_status = c.is_tool_btn_clicked(2)
        if current_status == [1] and last_status == [0]:
            print("检测到红色按钮按下，开始发送角度数据...")
            break
        last_status = current_status
        time.sleep(0.05)


def processing_data(angle):
    angle = [float(i) for i in angle]
    angle[1] *= -1

    gripper_angle = angle.pop(-1)
    angle.append((gripper_angle - 0.08) / (-95.27 - 0.08) * (-123.13 + 1.23) - 1.23)

    joint_limits = [
        (-160, 160),
        (-80, 100),
        (-90, 75),
        (-160, 160),
        (-80, 80),
        (-150, 150),
        (-115, 0),
    ]

    for i in range(min(len(angle), len(joint_limits))):
        low, high = joint_limits[i]
        angle[i] = max(min(angle[i], high), low)

    return angle


class AngleTransfer:
    def __init__(self, myarm_c, myarm_m, speed=100):
        self.c = myarm_c
        self.m = myarm_m
        self.speed = speed
        self.running = False
        self.angle = []

        # 数据采集相关
        self.data_buffer = []
        self.count = 1
        self.last_time = time.time()
        current_time = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())
        self.xls_name = f'Data_{current_time}.xlsx'
        template_file = "Data.xlsx"
        if os.path.exists(template_file):
            shutil.copy(template_file, self.xls_name)
            print(f"已创建新表格: {self.xls_name}")
        else:
            print("未找到模板文件，将创建新 Excel 文件")
            wb = Workbook()
            ws = wb.active
            ws.append(["编号", "角度", "间隔(ms)"])
            wb.save(self.xls_name)

    def start(self):
        self.running = True
        threading.Thread(target=self._read_from_c).start()
        threading.Thread(target=self._send_to_m).start()
        threading.Thread(target=self._log_data).start()

    def _read_from_c(self):
        while self.running:
            angle = self.c.get_joints_angle()
            if angle:
                self.angle = angle
            time.sleep(0.05)

    def _send_to_m(self):
        while self.running:
            if self.angle:
                try:
                    processed = processing_data(self.angle)
                    processed[2] = processed[2] + 17
                    self.m.set_joints_angle(processed, self.speed)
                except MyArmDataException:
                    pass
            time.sleep(0.05)

    def _log_data(self):
        print("开始记录MyArmM角度数据...")
        while self.running:
            try:
                current_time = time.time()
                interval = (current_time - self.last_time) * 1000
                self.last_time = current_time

                left_angles = self.m.get_joints_angle()

                self.data_buffer.append([
                    self.count,
                    str(left_angles),
                    interval
                ])
                # print(f"采集第 {self.count} 条数据，间隔: {interval:.2f} ms")
                self.count += 1

                if len(self.data_buffer) >= 50:
                    wb = load_workbook(self.xls_name)
                    ws = wb.active
                    for row in self.data_buffer:
                        ws.append(row)
                    wb.save(self.xls_name)
                    wb.close()
                    print(f"已写入 {len(self.data_buffer)} 条数据到 Excel")
                    self.data_buffer.clear()

            except Exception as e:
                print("写入失败:", e)
                print(traceback.format_exc())
            time.sleep(0.01)


def main():
    # port_dict = get_port()
    # port_c = input("input myArm C port: ")
    # c_port = port_dict[port_c]
    # port_m = input("input myArm M port: ")
    # m_port = port_dict[port_m]

    myarm_c = MyArmC("COM43", debug=False)
    myarm_m = MyArmM("COM18", 1000000, debug=False)

    wait_for_button_press(myarm_c)

    controller = AngleTransfer(myarm_c, myarm_m)
    controller.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("已停止控制")
        controller.running = False


if __name__ == "__main__":
    main()
