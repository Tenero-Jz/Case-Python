import threading
import pandas as pd
import time
import os
from pymycobot import Mercury
from exoskeleton_api import Exoskeleton
from openpyxl import load_workbook

os.system("sudo chmod 777 /dev/ttyACM*")
# 初始化设备
obj = Exoskeleton(port="/dev/ttyACM4")

# 生成唯一的 Excel 文件名
timestamp = time.strftime("%Y%m%d_%H%M%S")
file_name = f"arm_data_{timestamp}.xlsx"

# 定义列名
columns = ["Joint1", "Joint2", "Joint3", "Joint4", "Joint5", "Joint6", "Joint7", "Button1", "Button2", "Arm"]

# 创建 Excel（如果文件不存在，则写入标题）
if not os.path.exists(file_name):
    df = pd.DataFrame(columns=columns)
    df.to_excel(file_name, index=False)  # 写入标题，不写入数据

# 定义数据缓存
data_buffer = []
buffer_lock = threading.Lock()  # 线程锁，防止多个线程同时写入冲突


# 限制关节范围
def jointlimit(angles):
    max_vals = [165, 120, 175, 0, 175, 180, 175]
    min_vals = [-165, -50, -175, -175, -175, 60, -175]
    for i in range(7):
        if angles[i] > max_vals[i]:
            angles[i] = max_vals[i]
        if angles[i] < min_vals[i]:
            angles[i] = min_vals[i]


# 采集数据并存入缓存，每 50 条写入一次 Excel
def control_arm(arm):
    global data_buffer
    while True:
        try:
            # 获取机械臂数据
            arm_data = obj.get_arm_data(arm)
            print(f"{'Left' if arm == 1 else 'Right'}: {arm_data}")

            # 提取所需数据
            mercury_list = [
                arm_data[0], -arm_data[1], arm_data[2], -arm_data[3], arm_data[4],
                135 + arm_data[5], arm_data[6] - 30
            ]
            jointlimit(mercury_list)

            button1 = arm_data[9]  # 第10个值（索引9）
            button2 = arm_data[10]  # 第11个值（索引10）

            # 组织新数据
            new_data = mercury_list + [button1, button2, "Left" if arm == 1 else "Right"]

            # 线程锁，确保多线程安全
            with buffer_lock:
                data_buffer.append(new_data)

                # 每 50 条数据写入一次 Excel
                if len(data_buffer) >= 50:
                    df = pd.DataFrame(data_buffer, columns=columns)

                    # 获取当前已有行数
                    if os.path.exists(file_name):
                        wb = load_workbook(file_name)
                        sheet = wb.active
                        startrow = sheet.max_row  # 获取当前 Excel 最后一行
                    else:
                        startrow = 0  # 如果文件不存在，则从头开始

                    # 追加到 Excel
                    with pd.ExcelWriter(file_name, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                        df.to_excel(writer, index=False, header=False, startrow=startrow)

                    print(f"写入 50 条数据到 {file_name}（从第 {startrow+1} 行开始）")
                    data_buffer = []  # 清空缓存

            time.sleep(0.01)  # 适当延时，避免过多写入

        except Exception as e:
            print(f"错误: {e}")
            pass


# 启动线程（去掉 daemon=True）
thread1 = threading.Thread(target=control_arm, args=(1,))
thread2 = threading.Thread(target=control_arm, args=(2,))
thread1.start()
thread2.start()

# **主线程等待子线程结束**
thread1.join()
thread2.join()
